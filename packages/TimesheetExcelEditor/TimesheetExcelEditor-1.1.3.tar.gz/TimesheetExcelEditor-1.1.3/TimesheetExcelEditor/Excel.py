import openpyxl
import re
import os
from . import config
from datetime import date as dt, timedelta
from openpyxl.utils import get_column_letter
from openpyxl import formatting
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill


class Excel:

    def __init__(self):
        self.filename = None
        self.filePath = None
        self.wb = None
        self.headers = {}
        self.missing = []

    def load(self, directory, name):
        self.filename = name
        self.filePath = os.path.join(directory, name)
        self.wb = openpyxl.load_workbook(self.filePath)

    def createWorkbook(self, year, add):
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb['Sheet'])
        self.__createSheet(year)
        for times in range(1, add+1):
            newYear = str(int(year)+times)
            self.__createSheet(newYear)
        return self.wb

    def __createSheet(self, year):
        self.wb.create_sheet(year)
        ws = self.wb[year]
        ws['A2'] = "RIEPILOGO " + str(int(year) - 1)
        ws['A4'] = "TOTALE ORE"
        ws['A5'] = "Attività istituzionale"
        ws['U2'] = "Previsione"
        number = 1
        cursiveFont = Font(italic=True)
        for acRow in range(6, 6+8):
            column = 'A'
            ws[column+str(acRow)] = "Progetto " + str(number)
            ws[column+str(acRow)].font = cursiveFont
            number += 1
        ws['A15'] = "TOTALE"
        months = ['G', 'F', 'M', 'A', 'M', 'G', 'L', 'A', 'S', 'O', 'N', 'D', 'TOT']
        monthsIndex = 0
        row = 21
        increment = 0
        boldFont = Font(bold=True)
        border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"),
                        bottom=Side(border_style="thin"))
        for r in range(3, 16):
            ws['A'+str(r)].border = border
        for col in range(2, 15):
            for rows in range(3, 14):
                if col != 14:
                    if rows == 3:
                        ws.cell(row=rows, column=col).value = months[monthsIndex]
                        ws.cell(row=rows, column=col).border = border
                        ws.cell(row=rows, column=col + 19).value = months[monthsIndex]
                        ws.cell(row=rows, column=col + 19).border = border
                        monthsIndex += 1
                    elif rows == 4:
                        ws.cell(row=rows, column=col).value = 125
                        ws.cell(row=rows, column=col).border = border
                        cell = ws.cell(row=rows, column=col).coordinate
                        ws.cell(row=rows, column=col + 19).value = "=" + cell
                        ws.cell(row=rows, column=col + 19).border = border
                    else:
                        column = 'AG'
                        if increment == 0:
                            ws.cell(row=rows, column=col).value = "="+column+str(row)
                            ws.cell(row=rows, column=col).border = border
                            c = ws.cell(row=rows, column=col).coordinate
                            ws.cell(row=rows, column=col + 19).value = "=" + c
                            ws.cell(row=rows, column=col + 19).border = border
                            increment += 1
                        else:
                            tempRow = row
                            ws.cell(row=rows, column=col).value = "=" + column + str(tempRow+increment)
                            ws.cell(row=rows, column=col).border = border
                            increment = (increment + 1) % 9
                            c = ws.cell(row=rows, column=col).coordinate
                            ws.cell(row=rows, column=col + 19).value = "=" + c
                            ws.cell(row=rows, column=col + 19).border = border
                else:
                    if rows == 3:
                        ws.cell(row=rows, column=col).value = months[monthsIndex]
                        ws.cell(row=rows, column=col).font = boldFont
                        ws.cell(row=rows, column=col).border = border
                        ws.cell(row=rows, column=col + 19).value = months[monthsIndex]
                        ws.cell(row=rows, column=col + 19).font = boldFont
                        ws.cell(row=rows, column=col + 19).border = border
                    else:
                        start = ws.cell(row=rows, column=2).coordinate
                        end = ws.cell(row=rows, column=col-1).coordinate
                        pStart = ws.cell(row=rows, column=21).coordinate
                        pEnd = ws.cell(row=rows, column=col - 1+19).coordinate
                        ws.cell(row=rows, column=col).value = "=SUM("+start+":"+end+")"
                        ws.cell(row=rows, column=col).font = boldFont
                        ws.cell(row=rows, column=col).border = border
                        ws.cell(row=rows, column=col+19).value = "=SUM(" + pStart + ":" + pEnd + ")"
                        ws.cell(row=rows, column=col + 19).font = boldFont
                        ws.cell(row=rows, column=col + 19).border = border
            row += 14
            ws['A15'].font = boldFont
            ws['A15'].border = border
            tot = ws.cell(row=4, column=col).coordinate
            start = ws.cell(row=5, column=col).coordinate
            end = ws.cell(row=14, column=col).coordinate
            ws.cell(row=15, column=col).value = "="+tot+"-SUM(" + start + ":" + end + ")"
            ws.cell(row=15, column=col).font = boldFont
            ws.cell(row=15, column=col).border = border
            pTot = ws.cell(row=4, column=col+19).coordinate
            pStart = ws.cell(row=5, column=col+19).coordinate
            pEnd = ws.cell(row=14, column=col+19).coordinate
            ws.cell(row=15, column=col + 19).value = "="+pTot+"-SUM(" + pStart + ":" + pEnd + ")"
            ws.cell(row=15, column=col + 19).font = boldFont
            ws.cell(row=15, column=col + 19).border = border
        ws['A17'] = "DETTAGLIO MENSILE"
        monthsDic = {"01": "GENNAIO", "02": "FEBBRAIO", "03": "MARZO", "04": "APRILE", "05": "MAGGIO", "06": "GIUGNO",
                     "07": "LUGLIO", "08": "AGOSTO", "09": "SETTEMBRE", "10": "OTTOBRE", "11": "NOVEMBRE",
                     "12": "DICEMBRE"}
        rowStart = 20
        for keys in monthsDic:
            column = 'A'
            ws.merge_cells('B'+str(rowStart-2)+':'+'D'+str(rowStart-2))
            ws['B'+str(rowStart-2)] = monthsDic[keys]
            ws['A'+str(rowStart-1)].border = border
            monthStart, monthEnd = self.__getMonthDays(int(year), keys, ws)
            for c in range(monthStart[1], monthEnd+1):
                ws.cell(row=rowStart-1, column=c).value = c - 1
                ws.cell(row=rowStart-1, column=c).border = border
                ws.cell(row=rowStart - 1, column=c).alignment = Alignment(horizontal='center')
                day = dt(int(year), int(keys), c - 1).strftime("%A")
                if day == "Saturday" or day == "Sunday":
                    color = None
                    if int(year) % 2 == 0:
                        color = PatternFill(start_color="FFA6A6A6", end_color="FFA6A6A6", fill_type="solid")
                    else:
                        color = PatternFill(start_color="FF8EB4E3", end_color="FF8EB4E3", fill_type="solid")
                    for row in range(rowStart - 1, rowStart + 11):
                        ws.cell(row=row, column=c).fill = color
            ws['AG'+str(rowStart-1)] = "TOT"
            ws['AG'+str(rowStart-1)].alignment = Alignment(horizontal='center')
            number = 6
            terminate = rowStart + 11
            first = True
            while rowStart < terminate:
                if first:
                    cellRange = 'B'+str(rowStart+10)+':AG'+str(rowStart+10)
                    lightRed = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")
                    ws.conditional_formatting.add(cellRange, formatting.rule.CellIsRule(operator='lessThan',
                                                                                        formula=['0'], fill=lightRed))
                    for c in range(2, 33):
                        totHours = ws.cell(row=rowStart, column=c).coordinate
                        startS = ws.cell(row=rowStart+1, column=c).coordinate
                        endS = ws.cell(row=rowStart+9, column=c).coordinate
                        ws.cell(row=rowStart+10, column=c).value = "=" + totHours + "-SUM(" + startS + ":" + endS + ")"
                    first = False
                if terminate-1 - rowStart == 10:
                    ws[column + str(rowStart)] = "TOTALE ORE"
                elif terminate-1 - rowStart == 9:
                    ws[column + str(rowStart)] = "Attività istituzionale"
                elif terminate-1 - rowStart != 0:
                    ws[column + str(rowStart)] = "="+column + str(number)
                    number += 1
                else:
                    ws[column + str(rowStart)] = "RESIDUO"
                ws['AG' + str(rowStart)] = '=SUM(' + 'B' + str(rowStart) + ':' + 'AF' + str(rowStart) + ')'
                rowStart += 1
            rowStart += 3
        ws.column_dimensions['A'].width = 20
        for c in range(2, 34):
            ws.column_dimensions[get_column_letter(c)].width = 4
        ws.column_dimensions['N'].width = 5
        ws.column_dimensions['AG'].width = 5
        return ws

    def getFileName(self):
        return self.filename

    def getPath(self):
        return self.filePath

    def checkMissing(self):
        return len(self.missing)

    def getMissing(self):
        message = "Anno/i "
        for y in self.missing:
            message += y + ", "
        message += "non trovato/i"
        self.missing.clear()
        return message

    def save(self, path):
        self.wb.save(path)
        config.log.AppendText("File salvato in "+path+"\n")

    def getHeaders(self):
        ws = self.wb.active
        self.headers["A1"] = "Attività istituzionale"
        for row in ws.iter_rows(max_col=1):
            for cell in row:
                m = re.match("Progetto \d+", str(cell.value))
                if m is not None:
                    self.headers[cell.coordinate] = m.group()
        return self.headers

    def __checkSheet(self, year):
        sheets = self.wb.sheetnames
        present = False
        for s in sheets:
            if s == year:
                present = True
        if not present:
            if year not in self.missing:
                self.missing.append(year)
        return present

    def __getMonthDays(self, year, m, ws):
        monthsDic = {"01": "GENNAIO", "02": "FEBBRAIO", "03": "MARZO", "04": "APRILE", "05": "MAGGIO", "06": "GIUGNO",
                     "07": "LUGLIO", "08": "AGOSTO", "09": "SETTEMBRE", "10": "OTTOBRE", "11": "NOVEMBRE",
                     "12": "DICEMBRE"}
        for months in ws.iter_rows(min_col=2, max_col=2):
            for cell in months:
                if cell.value is not None:
                    value = str(cell.value)
                    for keys in monthsDic:
                        if m == keys and re.search(monthsDic[keys], value, re.IGNORECASE):
                            start = (cell.row + 1, cell.column)
                            n = int(keys)
                            if n == 2:
                                if year % 4 == 0 and year % 100 != 0 or year % 400 == 0:
                                    end = cell.column + 28
                                else:
                                    end = cell.column + 27
                            elif n == 1 or n == 3 or n == 5 or n == 7 or n == 8 or n == 10 or n == 12:
                                end = cell.column + 30
                            else:
                                end = cell.column + 29
                            return start, end

    def modify(self, activity, data, nore, multiplier):
        numOre = nore
        day, month, year = data.split("/")
        date = dt(int(year), int(month), int(day))
        if self.__checkSheet(year):
            ws = self.wb[year]
            start, end = self.__getMonthDays(int(year), month, ws)
            for days in ws.iter_rows(min_row=start[0], max_row=start[0], min_col=start[1], max_col=end):
                for cell in days:
                    if cell.value == int(day):
                        for activities in ws.iter_rows(min_row=cell.row + 1, max_col=1):
                            for cell2 in activities:
                                value = None
                                for keys in self.headers:
                                    if self.headers[keys] == activity:
                                        value = "=" + keys
                                if cell2.value == value or cell2.value == activity:
                                    point = cell.column_letter + str(cell2.row)
                                    if ws[point].value is not None:
                                        ws[point] = ws[point].value + round(numOre * multiplier)
                                    else:
                                        ws[point] = round(numOre * multiplier)
                                    config.log.AppendText("Aggiunte "+str(round(numOre * multiplier)) +
                                                          " ore di "+activity+" alla data "+data + ".\n")
                                    if ws[point].value > 7:
                                        if not date.month == date.day == 1 and date.year == int(year):
                                            plus = ws[point].value - 7
                                            config.log.AppendText("Limite 7 ore superato, "+str(plus) +
                                                                  " ore in eccesso " +
                                                                  "ridistribuite nei giorni precedenti.\n")
                                            ws[point] = 7
                                            new_row = cell2.row
                                            new_col = cell.column - 1
                                            new_date = date - timedelta(days=1)
                                            while plus > 0 and (new_date.year == date.year):
                                                if new_date.month == date.month - 1:
                                                    null, new_col = self.__getMonthDays(int(year),
                                                                                        new_date.strftime("%m"), ws)
                                                    new_row = null[0] + (cell2.row - start[0])
                                                    date = new_date
                                                if new_date.weekday() != 6:
                                                    if ws.cell(new_row, new_col).value is not None:
                                                        ws.cell(new_row, new_col).value += 1
                                                    else:
                                                        ws.cell(new_row, new_col).value = 1
                                                    plus -= 1
                                                    if ws.cell(new_row, new_col).value == 7:
                                                        if not new_date.month == new_date.day == 1:
                                                            new_col -= 1
                                                            new_date = new_date - timedelta(days=1)
                                                else:
                                                    new_col -= 1
                                                    new_date = new_date - timedelta(days=1)
                                            if plus > 0:
                                                if ws.cell(new_row, new_col).value is not None:
                                                    ws.cell(new_row, new_col).value += plus
                                                else:
                                                    ws.cell(new_row, new_col).value = plus
                                    return True
