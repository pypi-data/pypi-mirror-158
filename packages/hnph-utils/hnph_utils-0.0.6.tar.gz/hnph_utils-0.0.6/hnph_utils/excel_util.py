import xlrd
import openpyxl


class ExcelUtil:
    @staticmethod
    def load_excel_data_xls(file_path: str, cols: int, row_begin=2) -> list:
        """
        加载 Excel xls 表格数据
        :param file_path: 文件路径
        :param cols: 总列数
        :param row_begin: 有效数据开始行数，默认第 2 行
        :return:
        """
        datas = []
        workbook = xlrd.open_workbook(file_path, encoding_override='utf-8', ragged_rows=True)
        for sheet_name in workbook.sheet_names():
            sheet = workbook.sheet_by_name(sheet_name)

            if sheet.nrows > 0 and sheet.ncols > 0:
                for i in range(sheet.nrows + row_begin - 3):
                    row_data = []
                    for j in range(cols):
                        try:
                            value = sheet.cell(i + 1, j).value
                        except:
                            value = None
                        if any(row_data):
                            row_data.append(value)
                    datas.append(row_data)
        return datas

    @staticmethod
    def load_excel_data_xlsx(file_path: str, cols: int, row_begin=2) -> list:
        """
        加载 Excel xlsx 表格数据
        :param file_path: 文件路径
        :param cols: 总列数
        :param row_begin: 有效数据开始行数，默认第 2 行
        :return:
        """
        datas = []
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(min_row=row_begin):
                row_data = []
                for i in range(cols):
                    row_data.append(row[i].value)
                if any(row_data):
                    datas.append(row_data)

        return datas

    @staticmethod
    def load_excel_data(file_path: str, cols: int, row_begin=2) -> list:
        """
        加载 Excel 表格数据
        :param file_path: Excel文件绝对路径
        :param cols: 总列数
        :param row_begin: 有效数据开始行数，默认第 2 行
        :return:
        """
        if file_path.upper().endswith('.XLS'):
            return ExcelUtil.load_excel_data_xls(file_path, cols, row_begin)
        else:
            return ExcelUtil.load_excel_data_xlsx(file_path, cols, row_begin)
