import time
import traceback
from functools import wraps, reduce
from XLRP.drawplot import DrawPlot
from pathlib import Path
import os
from XLRP.xlconf import XlConf
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from XLRP.xl_ddt import *
from luckylog import luckylog
from luckylog.luckylog import Logger
import sys
import threading


_test_result = {"sys": '', "model": [], 'step': []}
_error = {}
_passed_count = 0
_failed_count = 0
_error_count = 0
CACHE_PATH = Path().cwd() / '.xlrp_cache'
CACHE_PATH.mkdir(exist_ok=True)
_XL_LOCK = threading.RLock()
file_mark = None
log_file = None
_old_model = None
_step_count = 0
_current_models = []


def _log_path():
    path_dir = CACHE_PATH / 'logs'
    path_dir.mkdir(exist_ok=True)
    file = path_dir / f'xlrp{file_mark}.log'
    file.touch(exist_ok=True)
    file_path = os.fspath(file)
    return file_path

def _write_log(file_path, msg):
    with open(file_path, 'a', encoding='utf8') as f:
        f.writelines(msg)
        f.write('\n')
        f.close()

# log_file = _log_path()
# luckylog.path = log_file
# luckylog.module = 'success, error, warning, tip'


class SysName:
    """
    获取系统名称的类
    """
    def __init__(self, sys_name):
        self.sys_name = sys_name

    def __enter__(self):
        pass

    def __exit__(self, exc_type, exc_val, exc_tb):
        _test_result["sys"] = self.sys_name

    def __call__(self, cls):
        _test_result['sys'] = self.sys_name

        @wraps(cls)
        def decorator(*args, **kwargs):
            res = cls(*args, **kwargs)
            return res
        return decorator

class ModelName:
    """
    获取模块名称类
    """
    def __init__(self, model_name):
        self.model = model_name

    def __enter__(self):
        global _old_model
        _old_model = self.model

    def __exit__(self, exc_type, exc_val, exc_tb):
        ...

    def __call__(self, func):
        @wraps(func)
        def decorator(*args, **kwargs):
            global _old_model
            _old_model = self.model
            # _test_result["model"].append(self.model)
            res = func(*args, **kwargs)
            return res
        return decorator

class StepName:
    """
    获取step步骤(用例)的类
    """
    def __init__(self, step_name):
        self.step = step_name

    def __enter__(self):
        self.start_time = time.time()

    def __exit__(self, exc_type, exc_val, exc_tb):
        global _failed_count, _passed_count, _error_count, _test_result, _old_model
        if any([exc_type, exc_val, exc_tb]):
            # error_str = str([''.join(x) for x in traceback.format_exception(exc_type, exc_val, exc_tb)][0])
            error_str = reduce(lambda x, y: x + y, traceback.format_exception(exc_type, exc_val, exc_tb))
            _error[self.step] = error_str
            if exc_type is AssertionError:
                case_status = False
                Logger.error(f'{self.step}----Failed')
                _failed_count += 1
            else:
                case_status = 'Error'
                Logger.error(f'{self.step}----Error')
                _error_count += 1
            e = traceback.format_exception(exc_type, exc_val, exc_tb)
            _write_log(log_file, e)
            if type(e) is list:
                for msg in e:
                    print(msg, end='')
            else:
                print(e)
            print()
        else:
            _passed_count += 1
            case_status = True
            error_str = None
            Logger.success(f'{self.step}---Passed')
        stop_time = time.time()
        run_time = round(stop_time - self.start_time, 3)
        _test_result["step"].append({self.step: {'run_time': run_time, 'status': case_status, 'msg': error_str}})
        _test_result['model'].append(_old_model)

    def __call__(self, func):
        step = self.step

        @wraps(func)
        def decorator(*args, **kwargs):
            global _passed_count, _failed_count, _error_count, _test_result, _old_model
            start_time = time.time()
            try:
                res = func(*args, **kwargs)
                if type(res) is bool:
                    case_status = res
                    if case_status is False:
                        log = f'{step}---Failed'
                        _failed_count += 1
                        Logger.error(log)
                        _write_log(log_file, f'The custom assertion failed, function {func.__name__} return {res}')
                        print(f'The custom assertion failed, function {func.__name__} return {res}\n')
                    else:
                        log = f'{step}----Passed'
                        Logger.success(log)
                        _passed_count += 1
                else:
                    case_status = True
                    log = f'{step}----Passed'
                    Logger.success(log)
                    _passed_count += 1
                error_str = None
            except Exception as e:
                case_status = False
                res = None
                # error_str = str([''.join(x) for x in traceback.format_exc()][0])
                error_str = reduce(lambda x, y: x + y, traceback.format_exc())
                exc_type, exc_val, exc_tb = sys.exc_info()
                if exc_type is AssertionError:
                    log = f'{step}----Failed'
                    _failed_count += 1
                else:
                    log = f'{step}----Error'
                    _error_count += 1
                Logger.error(log)
                e = traceback.format_exc()
                _write_log(log_file, e)
                if type(e) is list:
                    for msg in e:
                        print(msg)
                else:
                    print(e)
                print()
            stop_time = time.time()
            run_time = round(stop_time - start_time, 3)
            _test_result['model'].append(_old_model)
            _test_result["step"].append({step: {"run_time": run_time, "status": case_status, 'msg': error_str}})
            return res
        return decorator

class _ProcessingData:
    def __init__(self, data_list: dict):
        self.data_list = data_list
        self.model_list = self.data_list['model']
        self.filter_models = []
        [self.filter_models.append(x) for x in self.model_list if x not in self.filter_models]
        self.steps = self.data_list['step']
        self.zips_data = list(zip(self.model_list, self.steps))

    def pie_data(self):
        """
        修整饼图数据
        :return:
        """
        passed = 0
        failed = 0
        error = 0
        for dc in self.steps:
            status = list(dc.values())[0]['status']
            if status is True:
                passed += 1
            elif status == 'Error':
                error += 1
            elif status is False:
                failed += 1
            else:
                failed += 1
        return passed, failed, error

    def bar_data(self):
        """
        修整柱状图的测试数据
        :return:
        """
        dc_data = {}
        for model in self.filter_models:
            dc_data[model] = {'通过': 0, '失败': 0, '错误': 0}
            for case in self.zips_data:
                if model == case[0]:
                    status = list(case[1].values())[0]['status']
                    if status is True:
                        dc_data[model]['通过'] += 1
                    elif status is False:
                        dc_data[model]['失败'] += 1
                    elif status == 'Error':
                        dc_data[model]['错误'] += 1
                    else:
                        dc_data[model]['失败'] += 1
        return dc_data

    def plot_data(self):
        """
        修正绘制折线图的测试数据
        :return:
        """
        dc_data = {}
        for model in self.filter_models:
            dc_data[model] = {}
            for case in self.zips_data:
                if model == case[0]:
                    run_time = {list(case[1].keys())[0]: list(case[1].values())[0]["run_time"]}
                    dc_data[model].update(run_time)
        return dc_data

class Runner:
    def __init__(self, show_plot=False, file_mark=''):
        self.show_plot = show_plot
        self.save_path = CACHE_PATH / 'plot'
        self.save_path.mkdir(exist_ok=True)
        self.save_path = os.fspath(self.save_path)
        self.file_mark = file_mark

    def run(self, obj, param_iter=None):
        """
        运行单个用例的方法
        :param obj: 用例函数或者方法名
        :param param_iter: 参数，使用同parameter，需要列表或者元组形式，会循环传入到函数或者方法中
        :return: self
        """
        global file_mark, log_file
        if self.file_mark:
            file_mark = '_' + self.file_mark
        else:
            file_mark = self.file_mark
        log_file = _log_path()
        luckylog.path = log_file
        luckylog.module = 'success, error, warning, tip'
        start_time = time.strftime('%Y/%m/%d-%H:%M:%S')
        start_label = '=' * 30 + start_time + '  开始运行' + '=' * 30
        _write_log(file_path=log_file, msg=start_label)
        print('='*30 + ' 开始测试 ' + '='*30)
        print('-'*20 + f' 开始用例{obj.__name__} ' + '-'*20)
        _write_log(log_file, '-'*20 + f' 开始用例{obj.__name__} ' + '-'*20)
        if type(param_iter) not in (list, tuple):
            if param_iter is None:
                try:
                    obj()
                except:
                    ...
            else:
                raise ValueError("The parameters must be lists or tuples")
        else:
            for param in param_iter:
                print(f'\033[1;50;34mparams: {param}\033[0m')
                _write_log(log_file, f'params: {param}')
                try:
                    if type(param) in (list, tuple):
                        obj(*param)
                    elif type(param) is dict:
                        obj(**param)
                    else:
                        obj(param)
                except:
                    ...
        print('-'*20 + f' 结束用例{obj.__name__} ' + '-'*20 + '\n')
        print('='*30, f'Passed: {_passed_count} Failed: {_failed_count} Error: {_error_count}', '='*30)
        _write_log(log_file, '-'*20 + f' 结束用例{obj.__name__} ' + '-'*20 + '\n')
        _write_log(log_file, '='*30 + f'Passed: {_passed_count} Failed: {_failed_count} Error: {_error_count}' + '='*30)
        end_time = time.strftime('%Y/%m/%d-%H:%M:%S')
        end_label = '='*30 + end_time + '  结束run运行' + '=' * 30
        _write_log(log_file, end_label + '\n')
        return self

    def run_class(self, cls):
        global file_mark, log_file
        if self.file_mark:
            file_mark = '_' + self.file_mark
        else:
            file_mark = self.file_mark
        log_file = _log_path()
        luckylog.path = log_file
        luckylog.module = 'success, error, warning, tip'
        # print('user_func', USE_FUNC)
        start_time = time.strftime('%Y/%m/%d-%H:%M:%S')
        start_label = '=' * 30 + start_time + '  开始运行' + '=' * 30
        _write_log(file_path=log_file, msg=start_label)
        print('='*50 + ' 开始测试 ' + '='*50)
        func_names = [x.__name__ for x in FUNC_NAME]
        all_func = [x for x in cls.__dir__() if x.startswith('xl') and (x not in func_names)]
        user_func_key = [list(x.keys())[0] for x in USE_FUNC]
        runner_list = list(set(all_func).difference(set(user_func_key)))
        # runner_list.sort()
        # print(runner_list)
        [USE_FUNC.append({x: ()}) for x in runner_list]
        USE_FUNC.sort(key=lambda x: list(x.keys())[0])
        for data in USE_FUNC:
            func = list(data.keys())[0]
            param = list(data.values())[0]
            print('-'*20 + f' 开始用例{func} ' + '-'*20)
            _write_log(log_file, '-'*20 + f' 开始用例{func} ' + '-'*20)
            try:
                _write_log(log_file, f'params: {param}')
                print(f'\033[1;50;34mparams: {param}\033[0m')
                # func = list(data.keys())[0]
                # param = list(data.values())[0]
                if type(param) in (list, tuple):
                    eval(f'cls.{func}(*{param})')
                elif type(param) is dict:
                    eval(f'cls.{func}(**{param})')
                else:
                    eval(f'cls.{func}({param})')
            except:
                ...
            print('-'*20 + f' 结束用例{func} ' + '-'*20 + '\n')
            _write_log(log_file, '-'*20 + f' 结束用例{func} ' + '-'*20 + '\n')
        print('=' * 30, f'Passed: {_passed_count} Failed: {_failed_count} Error: {_error_count}', '=' * 30)
        _write_log(log_file, '=' * 30 + f'Passed: {_passed_count} Failed: {_failed_count} Error: {_error_count}' + '=' * 30)
        end_time = time.strftime('%Y/%m/%d-%H:%M:%S')
        end_label = '='*30 + end_time + '  结束class运行' + '=' * 30
        _write_log(log_file, end_label + '\n')
        return self

    def plot_save_excel(self, file_path=None, width_list=(700, 700, 700)):
        """
        将plot图片保存到excel的代码
        :param file_path: excel文件路径
        :param width_list: 图片的宽度(默认第三个数据，会在原基础上X2，改动在xlconf文件中)
        :return: self
        """
        sys_name = _test_result.get('sys')
        prd = _ProcessingData(_test_result)
        pie_data = prd.pie_data()
        bar_data = prd.bar_data()
        plot_data = prd.plot_data()
        dp = DrawPlot(sys_name=sys_name,file_mark=file_mark, save_path=self.save_path)
        pie_path = dp.draw_pie(通过=pie_data[0], 失败=pie_data[1], 错误=pie_data[2], show_plot=self.show_plot)
        bar_path = dp.draw_bar(**bar_data, show_plot=self.show_plot)
        plot_path = dp.draw_plot(**plot_data, show_plot=self.show_plot)
        pic_list = [pie_path, bar_path, plot_path]
        if file_path is not None:
            # dp.plot_to_excel(file_path=file_path, pic_list=pic_list, width_list=width_list, cells=cells)
            xc = XlConf(filepath=file_path, sheetname='TestReport', pic_list=pic_list, pic_width=width_list)
            xc.set_cell_size(create_new=True)
            del xc
            xc2 = XlConf(filepath=file_path, sheetname='TestReport', pic_list=pic_list, pic_width=width_list)
            xc2.save_plot()

    def save_default(self):
        """
        创建excel保存用例数据，并新建sheet存放报告图表
        :return:
        """
        global CACHE_PATH
        sys_name = _test_result.get('sys')
        excel_name = f'{sys_name}EXCEL报告{file_mark}.xlsx'
        report_path = CACHE_PATH / 'report'
        report_path.mkdir(exist_ok=True)
        model_list = _test_result['model']
        steps = _test_result['step']
        # excel样式设置
        wb = openpyxl.Workbook()
        ws = wb.create_sheet('TestCases')
        title_font = Font(size=16, name='微软雅黑', bold=True)
        label_font = Font(size=14, name='微软雅黑')
        content_font = Font(size=11, name='微软雅黑', bold=True)
        false_font = Font(size=11, name='微软雅黑', color='cc3333', bold=True)
        title_fill = PatternFill(fgColor='adc2eb', fill_type='solid')
        label_fill = PatternFill(fgColor='b3b3b3', fill_type='solid')
        # false_fill = PatternFill(fgColor='cc3333', fill_type='solid')
        cell_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.merge_cells('A1:F1')
        ws['A1'].value = f'{sys_name}系统运行用例情况表'
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws['A1'].alignment = cell_align
        labels = ['编号', '模块', '用例名称', '运行时间', '运行结果', '异常信息']
        # excel用例存放
        for index, label in enumerate(labels):                  # 设置表格第二行的标签
            label_cell = ws.cell(row=2, column=index + 1)
            label_cell.value = label
            label_cell.font = label_font
            label_cell.fill = label_fill
            label_cell.alignment = cell_align
        for index, case_data in enumerate(zip(model_list, steps)):
            model = case_data[0]
            step = case_data[1]
            case_name = list(step.keys())[0]
            runtime = step[case_name]['run_time']
            status_origin = step[case_name]['status']
            if status_origin is True:
                status = 'PASSED'
            elif status_origin is False:
                status = 'FAILED'
            else:
                status = 'ERROR'
            msg = step[case_name]['msg']
            data = [str(index + 1), model, case_name, runtime, status, msg]
            for col, param in enumerate(data):
                case_cell = ws.cell(row=index + 3, column=col + 1)
                case_cell.value = param
                if status != 'PASSED':
                    case_cell.font = false_font
                else:
                    case_cell.font = content_font
                case_cell.alignment = cell_align
                # if status is not True:
                #     case_cell.fill = false_fill
        sava_path = os.fspath(report_path) + '/' + excel_name
        wb.save(sava_path)
        wb.close()
        del wb
        wb2 = openpyxl.load_workbook(sava_path)
        ws2 = wb2['TestCases']
        try:
            wb2.remove(wb2['Sheet'])
        except:
            ...
        for col in range(1, len(labels) + 1):
            width = 25
            if col == len(labels):
                width = 50
            ws2.column_dimensions[get_column_letter(col)].width = width
        old_model = ''
        for row in range(3, ws2.max_row + 1):
            current_model = ws2.cell(row=row, column=2).value
            if old_model == current_model:
                ws2.merge_cells(f'B{row - 1}:B{row}')
            else:
                old_model = current_model
        wb2.save(sava_path)
        wb2.close()
        del wb2
        self.plot_save_excel(file_path=sava_path)

