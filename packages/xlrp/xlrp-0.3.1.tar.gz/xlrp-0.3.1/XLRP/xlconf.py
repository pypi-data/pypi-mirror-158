from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image, PILImage
import win32gui, win32print, win32con


class XlConf:
    def __init__(self, filepath, sheetname:str, pic_list=None, pic_width=(500, 500, 500)):
        self.filepath = filepath
        self.pic_list = pic_list
        self.pic_width = pic_width
        self.sheet = sheetname
        self.dpi = win32print.GetDeviceCaps(win32gui.GetDC(0), win32con.LOGPIXELSX)

    def set_cell_size(self, cells=('A1', 'B1', 'A2'), create_new=False):
        wb = load_workbook(self.filepath)
        if create_new is True:
            ws = wb.create_sheet(self.sheet)
        else:
            ws = wb[self.sheet]
        for index, pic in enumerate(self.pic_list):
            w, h = PILImage.open(pic).size
            size_ratio = self.pic_width[index] / w
            cell_width = (self.pic_width[index] / self.dpi * 25.4) / 1.8
            cell_col = cells[index][0]
            cell_row = cells[index][1]
            ws.row_dimensions[cell_row].height = h * size_ratio
            ws.column_dimensions[cell_col].width = cell_width
        wb.save(self.filepath)
        wb.close()
        return self

    def save_plot(self, cells=('A1', 'B1', 'A2:B2')):
        wb = load_workbook(self.filepath)
        ws = wb[self.sheet]
        for index, pic in enumerate(self.pic_list):
            w, h = PILImage.open(pic).size
            ratio = self.pic_width[index] / w
            img = Image(pic)
            if index == 2:
                img.width, img.height = w * ratio * 2, h * ratio * 2
                ws.merge_cells(cells[index])
                cell = cells[index].split(':')[0]
            else:
                img.width, img.height = w * ratio, h * ratio
                cell = cells[index]
            align = Alignment(horizontal='center', vertical='center')
            ws[cell].alignment = align
            ws.add_image(img, cell)
        wb.save(self.filepath)
        wb.close()




