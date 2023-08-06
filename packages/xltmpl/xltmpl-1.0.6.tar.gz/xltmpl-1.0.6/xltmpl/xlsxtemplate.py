# -*- coding: utf-8 -*-
import sys
from . import worksheet as _worksheet
sys.modules['openpyxl.worksheet.worksheet'] = _worksheet

import re

import openpyxl.workbook
import pandas as pd

from xltmpl.xltemplatefactory import XLSheetBehaviorFactory
from .utils import index_to_column_letter


class XlsxTemplate(openpyxl.workbook.Workbook, XLSheetBehaviorFactory):
    def __init__(self, tmpl_path, save_path=None, apply_styles=True, place_holder='{{}}'):
        super(XlsxTemplate, self).__init__()
        self.tmpl_path = tmpl_path
        self._copy_template_attr()

        self.save_path = save_path
        self.apply_styles = apply_styles
        self.place_holder = place_holder
        self.sheet_headers: dict = {}

    def _copy_template_attr(self):
        wb = openpyxl.load_workbook(
            self.tmpl_path,
            read_only=False,
            keep_vba=True,
            data_only=False,
            keep_links=True
        )
        for attr in dir(wb):
            try:
                if attr.startswith('__') and attr.endswith('__'):
                    continue
                setattr(self, attr, getattr(wb, attr))
            except Exception as e:
                pass

    def _set_header_row(self, sheet_name, row_num=1):
        row = self[sheet_name][row_num:row_num]
        header = {index_to_column_letter(idx)
                  if isinstance(cell.value, (int, float, bool)) or cell.value is None
                  else cell.value
                  : index_to_column_letter(idx)
                  for idx, cell in enumerate(row)
                  }
        self.sheet_headers[sheet_name] = header

    def ensure_sheet_exist(self, sheet_name):
        try:
            sheet: openpyxl.worksheet.worksheet.Worksheet
            if isinstance(sheet_name, int):
                if 0 < sheet_name <= len(self.sheetnames):
                    sheet = self[self.sheetnames[sheet_name-1]]
                    self.sheet_headers[sheet.title] = None
                else:
                    sheet_name = str(sheet_name)
                    self.create_sheet(sheet_name)
                    sheet = self[sheet_name]
                    self.sheet_headers[sheet_name] = None
            elif isinstance(sheet_name, str):
                if sheet_name in self.sheetnames:
                    sheet = self[sheet_name]
                else:
                    self.create_sheet(sheet_name)
                    self.sheetnames.append(sheet_name)
                    sheet = self[sheet_name]
            else:
                return self[self.sheetnames[0]]
            return sheet
        except Exception as e:
            return self[self.sheetnames[0]]

    def append_row(self, sheet_name, row, style_row: int = None):
        sheet = self.ensure_sheet_exist(sheet_name)
        sheet.append(row, apply_styles=self.apply_styles, place_holder=self.place_holder)

    def append_rows(self, sheet_name, rows, style_row: int = None):
        sheet = self.ensure_sheet_exist(sheet_name)
        for row in rows:
            sheet.append(row, apply_styles=self.apply_styles, place_holder=self.place_holder)

    def append_dataframe(self, sheet_name, df: pd.DataFrame, keep_headers=False, header_row: int = None, style_row: int = None):
        sheet = self.ensure_sheet_exist(sheet_name)
        if keep_headers:
            sheet.append(df.columns.tolist(), apply_styles=self.apply_styles, place_holder=self.place_holder)
            for idx, row in df.iterrows():
                sheet.append(row.values.tolist(), apply_styles=self.apply_styles, place_holder=self.place_holder)
        else:
            for idx, row in df.iterrows():
                if header_row is None:
                    sheet.append(row.values.tolist(), apply_styles=self.apply_styles, place_holder=self.place_holder)
                else:
                    dict_ = row.to_dict()
                    if isinstance(header_row, int):
                        if 1 <= header_row <= 1048576:
                            self._set_header_row(sheet.title, header_row)
                            dict_data = {}
                            for field_name in self.sheet_headers[sheet.title]:
                                column_letter = self.sheet_headers[sheet.title][field_name]
                                if field_name in dict_:
                                    dict_data[column_letter] = dict_[field_name]
                            sheet.append(dict_data, apply_styles=self.apply_styles, place_holder=self.place_holder)
                    else:
                        raise Exception('header_row must be int.')

    def append_value(self, sheet_name, value, style_row: int = None):
        sheet = self.ensure_sheet_exist(sheet_name)
        sheet.append([value, ], apply_styles=self.apply_styles, place_holder=self.place_holder)

    def append_dict(self, sheet_name, dict_: dict, header_row=None, style_row: int = None):
        sheet = self.ensure_sheet_exist(sheet_name)
        if header_row is not None and isinstance(header_row, int):
            if 1 <= header_row <= 1048576:
                self._set_header_row(sheet.title, header_row)
                row = []
                dict_data = {}
                for field_name in self.sheet_headers[sheet.title]:
                    column_letter = self.sheet_headers[sheet.title][field_name]
                    if field_name in dict_:
                        dict_data[column_letter] = dict_[field_name]
                        row.append(dict_[field_name])
                    else:
                        row.append(None)
                # self._append(sheet.title, row, style_row)
                sheet.append(dict_data, apply_styles=self.apply_styles, place_holder=self.place_holder)

    def append(self, sheet_name, row, style_row: int = None):
        self._append(sheet_name, row, style_row)

    def _append(self, sheet_name, iterable, style_row: int = None):
        sheet = self.ensure_sheet_exist(sheet_name)
        sheet.append(iterable, apply_styles=self.apply_styles, place_holder=self.place_holder)

    def reg_replace(self, sheet_name, col, header_row, reg_mapper: dict):
        """{'邮政': '邮政快递包裹', '申通': '申通快递', '东莞邮政京东快递'}"""
        start_row = header_row + 1
        sheet = self.ensure_sheet_exist(sheet_name)
        header = [str(cell.value).strip().replace('\n', '') if cell.value is not None else cell.column_letter for cell in sheet[header_row: header_row]]
        if col in header:
            col_num = header.index(col) + 1
            if start_row > sheet.max_row:
                return
            regs: list = list(reg_mapper.keys())
            regs.sort(key=lambda x: len(x), reverse=True)
            for row_num in range(start_row, sheet.max_row+1):
                for reg in regs:
                    reg: str
                    if not (reg.startswith('^') and reg.endswith('$')):
                        reg_reshape = '.*' + reg + '.*'
                    cell_value = sheet.cell(row_num, col_num).value
                    cell_value = str(cell_value) if cell_value is not None else ''
                    if re.compile(reg_reshape).fullmatch(cell_value):
                        sheet.cell(row_num, col_num).value = reg_mapper[reg]
                        break

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.save_path is not None:
            self.save(self.save_path)
        self.close()

