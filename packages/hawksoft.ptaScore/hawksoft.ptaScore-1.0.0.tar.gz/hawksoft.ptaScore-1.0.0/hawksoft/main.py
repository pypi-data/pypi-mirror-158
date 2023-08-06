from openpyxl import load_workbook,Workbook
import argparse
import os

def retrieve(no,name,sheet):
    for idx,row in enumerate(sheet.rows):
        if row[1].value == no and row[2].value == name:
            if row[4].value == "缺考":
                return 0
            else:
                return row[4].value
    return 0
def work(names,tests,output):
    scoresBook = Workbook()
    scoresSheet = scoresBook.worksheets[0]
    scoresSheet.append(["编号","学号","姓名","学院"])
    scoresStartRow = 2

    nameBook = load_workbook(names)
    nameSheet = nameBook.worksheets[0]
    nameStartRow = 2

    for idx,row in enumerate(nameSheet.rows):
        if idx < nameStartRow:
            continue 
        no = row[0].value
        name = row[1].value
        department = row[3].value
        scoresSheet[chr(ord('A')+0)+str(idx-nameStartRow+2)] = idx-nameStartRow+1 
        scoresSheet[chr(ord('A')+1)+str(idx-nameStartRow+2)] = no 
        scoresSheet[chr(ord('A')+2)+str(idx-nameStartRow+2)] = name
        scoresSheet[chr(ord('A')+3)+str(idx-nameStartRow+2)] = department
        print(f"\r {idx-nameStartRow+1} {no} {name} {department}",end="")

        files = sorted(os.listdir(tests))
         
        testBooks = []
        for idfile,file in enumerate(files):
            testBook = load_workbook(os.path.join(tests,file))
            testBooks.append(testBook)   

        for idfile,file in enumerate(files):
            testBook = testBooks[idfile] # load_workbook("./tests/"+file)
            testSheet = testBook.worksheets[0]
            score = retrieve(no,name,testSheet)
            test =file.split(".")[-2]
            #print(f"{test}: {score}",end=" ")
            scoresSheet[chr(ord('A')+4+idfile)+"1"] = file.split(".")[-2]
            scoresSheet[chr(ord('A')+4+idfile)+str(idx-nameStartRow+2)] =score
        #print("\n")
    print(f"\r result in {output}. byebye.")
    scoresBook.save(output)
    for book in testBooks:
        book.close()
    scoresBook.close()
    nameBook.close()

def main():
    parse = argparse.ArgumentParser(description="根据学生名单names，统计tests下的多次考试成绩，生成一个output文件。")
    parse.add_argument("-n","--names",type=str,required=True,help="学生名单文件名,必须是excel文件")
    parse.add_argument("-t","--tests",type=str,required=True,help="测试成绩所在的目录,下面的文件都是excel文件")
    parse.add_argument("-o","--outputfile",type=str,default="scores.xlsx",help="结果文件名,excel文件")
    args = parse.parse_args()
    names = os.path.abspath(args.names)
    print(names)
    tests = os.path.abspath(args.tests)
    print(tests)
    output = os.path.abspath(args.outputfile)
    print(output)
    work(names,tests,output)
    
if __name__== "__main__":
    main()