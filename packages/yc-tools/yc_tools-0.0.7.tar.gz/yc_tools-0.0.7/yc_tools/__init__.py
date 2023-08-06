import os
import time
from time import sleep
import openpyxl
import winsound
import logging
import logging.handlers
from logging import *
import threading
logger = logging.getLogger()
logger.setLevel(logging.ERROR)
rht = logging.handlers.TimedRotatingFileHandler("3152.log", 'D')
fmt = logging.Formatter("%(asctime)s %(message)s %(funcName)s %(lineno)s ", "%Y-%m-%d %H:%M:%S")
rht.setFormatter(fmt)
logger.addHandler(rht)
debug = logger.debug
info = logger.info
warning = logger.warn
error = logger.error
critical = logger.critical

def Creat_Thread(thread,*args1):
    thread_uart_read=threading.Thread(target=thread,args=args1)
    thread_uart_read.setDaemon(True)
    thread_uart_read.start()
def get_cmd_result(cmd):
    print(cmd)
    result = os.popen(cmd)
    res = result.read()
    for line in res.splitlines():
        print (line)
    time.sleep(0.1)
    return res
def xw_toExcel(data,excelname):  # xlsxwriter库储存数据到excel
    try:
        workbook = openpyxl.load_workbook(excelname)  # 打开工作簿
        worksheet1 = workbook.active  # 创建子表
    except Exception as a:
        print(a)
        workbook = openpyxl.Workbook(excelname)
        workbook.save(filename= excelname)  # 关闭表
        workbook.close()
        workbook = openpyxl.load_workbook(excelname)  # 打开工作簿
        worksheet1 = workbook.active
    rows = worksheet1.max_row
    j = 0
    for i in data:
        j+=1
        worksheet1.cell(row=rows+1,column = j).value = i
    workbook.save(filename= excelname)  # 关闭表
    workbook.close()
def xw_columnToExcel(data,excelname):  # xlsxwriter库储存数据到excel
    try:
        workbook = openpyxl.load_workbook(excelname)  # 打开工作簿
        worksheet1 = workbook.active  # 创建子表
    except Exception as a:
        print(a)
        workbook = openpyxl.Workbook(excelname)
        workbook.save(filename= excelname)  # 关闭表
        workbook.close()
        workbook = openpyxl.load_workbook(excelname)  # 打开工作簿
        worksheet1 = workbook.active
    columns = worksheet1.max_column
    j = 0
    for i in data:
        j+=1
        worksheet1.cell(row=j,column = columns+1).value = i
    workbook.save(filename= excelname)  # 关闭表
    workbook.close()
def e_pu():
    while('CPU Stopped' not in get_cmd_result("e pu")):
        time.sleep(0.05)
        continue
    return
def e_p():
    while('CPU Stopped' not in get_cmd_result("e p")):
        time.sleep(0.05)
        continue
    return
def e_cu():
    while('CPU Running' not in get_cmd_result("e cu")):
        continue
    return
def e_c():
    while('CPU Running' not in get_cmd_result("e c")):
        continue
    return
def e_ku():
    while('CPU B reseted' not in get_cmd_result("e k")):
        continue
    return
def e_k():
    while('cm0 reseted' not in get_cmd_result("e k")):
        continue
    return
def count_bits_2(value):
    count = 0
    while(value):
        value &= value - 1
        count += 1
    return count
def to_hex(val):
    return hex(val).replace("0x","")

def beep_on():
    duration = 1000  # millisecond
    freq = 440  # Hz
    winsound.Beep(freq, duration)
def set_environ(var,val):
    os.environ[var]=val
