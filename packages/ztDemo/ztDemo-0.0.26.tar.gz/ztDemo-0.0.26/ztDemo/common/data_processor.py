# -*- coding: utf-8 -*-
import re

import yaml
import os

from openpyxl import load_workbook, Workbook
from openpyxl.cell import MergedCell


def data_provider(test_data):
    def wrapper(func):
        setattr(func, "__data_Provider__", test_data)
        global index_len
        index_len = len(str(len(test_data)))
        return func
    return wrapper



def mk_test_name(name, value, index=0):
    index = "{0:0{1}d}".format(index+1, index_len)
    test_name = "{0}_{1}_{2}".format(name, index, str(value))
    return re.sub(r'\W|^(?=\d)', '_', test_name)


class ExcelDataProcessor:
    def __init__(self, file_path):
        self.file_path = file_path

        if not file_path.endswith('.xlsx'):
            raise TypeError("不是一个有效的EXCEL文件")

        if not os.path.exists(self.file_path):
            Workbook().save(self.file_path)
            print("源文件不存在，在路径{}新建它".format(self.file_path))
        else:
            print("源文件{}存在".format(self.file_path))

    def read_data(self, sheet_name):
        all_data = []
        wb = load_workbook(self.file_path)
        if sheet_name not in wb.sheetnames:
            raise NameError('''Excel Tab "{}" 在文件 "{}" 中不存在'''.format(sheet_name, self.file_path))
        sheet = wb[sheet_name]
        print(f"Title = {sheet.title}")
        for row in sheet.rows:
            for cell in row:
                if isinstance(cell, MergedCell):
                    # 如果是Merged Cell就跳过
                    continue
                all_data.append(f"{cell.column_letter}{cell.row} = {cell.value}")
        return all_data

    def write_data(self, sheet_name, column, value):
        wb = load_workbook(self.file_path)
        wb.active
        if sheet_name not in wb.sheetnames:
            raise NameError('''Excel Tab "{}" 在文件 "{}" 中不存在'''.format(sheet_name, self.file_path))
        try:
            wb[sheet_name][column] = value
            wb.save(self.file_path)
        except:
            raise Exception("写入失败")



class YamlDataProcessor:
    def __init__(self, file_path):
        self.file_path = file_path

        if not file_path.endswith('.yaml'):
            raise TypeError("不是一个有效的YAML文件")

        if not os.path.exists(self.file_path):
            with open(self.file_path, 'w'):
                print("源文件不存在，在路径{}新建它".format(self.file_path))
        else:
            print("源文件{}存在".format(self.file_path))

    def read_yaml(self):
        with open(self.file_path, 'r') as stream:
            try:
                return yaml.safe_load(stream)
            except yaml.YAMLError as exc:
                print(exc)